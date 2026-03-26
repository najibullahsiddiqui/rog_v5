from __future__ import annotations

import io
import json
import logging
from pathlib import Path

from fastapi import APIRouter, HTTPException, Query, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook

from app.schemas import (
    CategoryPayload,
    CategorySynonymPayload,
    ChatSessionNotePayload,
    DataSourceCreatePayload,
    DataSourceStatusPayload,
    ExpertAnswerPayload,
    JsonConvertPayload,
    QnaPairPayload,
    TrainBotCategoryRefreshPayload,
    TrainBotPromoteExpertPayload,
    TrainBotPromoteQnaPayload,
    TrainBotReindexPayload,
    TrainBotResolveWrongAnswerPayload,
    WrongAnswerClassifyPayload,
    WrongAnswerConvertPayload,
    WrongAnswerResolvePayload,
    DecisionTreePayload,
)
from app.services import AnalyticsService, CategoriesService, ExpertAnswersService
from app.repositories import AdminRepository
from app.core.text_utils import normalize_question_text


router = APIRouter(tags=["admin"])
admin_repository = AdminRepository()
analytics_service = AnalyticsService(admin_repository)
expert_answers_service = ExpertAnswersService(admin_repository)
categories_service = CategoriesService()

BASE_DIR = Path(__file__).resolve().parent.parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
logger = logging.getLogger(__name__)


@router.get("/admin", response_class=HTMLResponse)
def admin_dashboard(request: Request):
    return templates.TemplateResponse(
        "admin_dashboard.html",
        {"request": request},
    )


@router.get("/api/admin/summary")
def get_summary():
    return analytics_service.summary()


@router.get("/admin/analytics", response_class=HTMLResponse)
def admin_analytics(request: Request):
    return templates.TemplateResponse(
        "admin_analytics.html",
        {"request": request},
    )


@router.get("/api/admin/dashboard-summary")
def get_dashboard_summary():
    return analytics_service.dashboard_summary()


@router.get("/api/admin/analytics")
def get_analytics(range_days: int = Query(default=30, ge=1, le=365)):
    return analytics_service.analytics_breakdown(range_days=range_days)


@router.get("/api/admin/unresolved")
def get_unresolved(
    category: str | None = Query(default=None),
    status: str = Query(default="open"),
):
    norm = categories_service.normalize(category) if category else None
    return {"items": admin_repository.list_unresolved(norm, status)}


@router.get("/api/admin/feedback")
def get_feedback(category: str | None = Query(default=None)):
    norm = categories_service.normalize(category) if category else None
    return {"items": admin_repository.list_feedback(norm)}


@router.get("/api/admin/wrong-answer-reports")
def list_wrong_answer_reports(
    status: str = Query(default="open"),
    limit: int = Query(default=200, ge=1, le=1000),
):
    return {"items": admin_repository.list_wrong_answer_reports(status=status, limit=limit)}


@router.post("/api/admin/wrong-answer-reports/{report_id}/classify")
def classify_wrong_answer_report(report_id: int, payload: WrongAnswerClassifyPayload):
    result = admin_repository.classify_wrong_answer_report(
        report_id=report_id,
        status=payload.status,
        assigned_to=payload.assigned_to,
        reason_code=payload.reason_code,
        severity=payload.severity,
        action_notes=payload.action_notes,
    )
    return {"ok": True, **result}


@router.post("/api/admin/wrong-answer-reports/{report_id}/resolve")
def resolve_wrong_answer_report(report_id: int, payload: WrongAnswerResolvePayload):
    result = admin_repository.resolve_wrong_answer_report(
        report_id=report_id,
        admin_action=payload.resolution_type,
        action_notes=payload.action_notes,
        resolution_type=payload.resolution_type,
    )
    return {"ok": True, **result}


@router.post("/api/admin/wrong-answer-reports/{report_id}/convert/expert")
def convert_wrong_report_to_expert(report_id: int, payload: WrongAnswerConvertPayload):
    category = categories_service.normalize(payload.category or "")
    if not category or not (payload.answer_text or "").strip():
        raise HTTPException(status_code=400, detail="category and answer_text are required")
    result = admin_repository.convert_wrong_answer_to_expert(
        report_id=report_id,
        category=category,
        expert_answer=str(payload.answer_text or "").strip(),
        source_note=payload.source_note,
    )
    return {"ok": True, **result}


@router.post("/api/admin/wrong-answer-reports/{report_id}/convert/qna")
def convert_wrong_report_to_qna(report_id: int, payload: WrongAnswerConvertPayload):
    if not (payload.answer_text or "").strip():
        raise HTTPException(status_code=400, detail="answer_text is required")
    result = admin_repository.convert_wrong_answer_to_qna(
        report_id=report_id,
        answer=str(payload.answer_text or "").strip(),
        category_code=categories_service.normalize(payload.category) if payload.category else None,
        source_note=payload.source_note,
    )
    return {"ok": True, **result}


@router.post("/api/admin/wrong-answer-reports/{report_id}/convert/category-fix")
def convert_wrong_report_to_category_fix(report_id: int, payload: WrongAnswerConvertPayload):
    result = admin_repository.convert_wrong_answer_to_category_fix(
        report_id=report_id,
        category_code=categories_service.normalize(payload.category) if payload.category else None,
        action_notes=payload.source_note,
    )
    return {"ok": True, **result}


@router.post("/api/admin/wrong-answer-reports/{report_id}/convert/source-issue")
def convert_wrong_report_to_source_issue(
    report_id: int,
    payload: WrongAnswerConvertPayload,
    data_source_id: int | None = Query(default=None),
):
    result = admin_repository.convert_wrong_answer_to_source_issue(
        report_id=report_id,
        data_source_id=data_source_id,
        action_notes=payload.source_note,
    )
    return {"ok": True, **result}


@router.get("/api/admin/chat-history/sessions")
def list_chat_history_sessions(
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    category: str | None = Query(default=None),
    response_mode: str | None = Query(default=None),
    feedback_status: str | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=1000),
):
    return {
        "items": admin_repository.list_chat_sessions(
            date_from=date_from,
            date_to=date_to,
            category_code=categories_service.normalize(category) if category else None,
            response_mode=response_mode,
            feedback_status=feedback_status,
            limit=limit,
        )
    }


@router.get("/api/admin/chat-history/sessions/{session_id}")
def get_chat_history_session_detail(session_id: int):
    item = admin_repository.get_chat_session_detail(session_id)
    if not item:
        raise HTTPException(status_code=404, detail="Session not found")
    return {"item": item}


@router.post("/api/admin/chat-history/sessions/{session_id}/note")
def update_chat_history_session_note(session_id: int, payload: ChatSessionNotePayload):
    ok = admin_repository.update_chat_session_note(session_id, payload.admin_note)
    if not ok:
        raise HTTPException(status_code=404, detail="Session not found")
    return {"ok": True}


@router.get("/api/admin/train-bot/queue")
def train_bot_queue(limit: int = Query(default=200, ge=1, le=500)):
    return {"items": admin_repository.get_train_bot_queue(limit=limit)}


@router.get("/api/admin/train-bot/jobs")
def train_bot_jobs(limit: int = Query(default=50, ge=1, le=200)):
    return {"items": admin_repository.list_training_jobs(limit=limit)}


@router.get("/api/admin/train-bot/audit")
def train_bot_audit(limit: int = Query(default=100, ge=1, le=500)):
    return {"items": admin_repository.list_audit_logs(limit=limit)}


@router.post("/api/admin/train-bot/actions/promote-expert")
def train_bot_promote_expert(payload: TrainBotPromoteExpertPayload):
    category = categories_service.normalize(payload.category)
    if not category:
        raise HTTPException(status_code=400, detail="Invalid category")
    result = admin_repository.promote_unresolved_to_expert(
        unresolved_query_id=payload.unresolved_query_id,
        category=category,
        expert_answer=payload.expert_answer,
        source_note=payload.source_note,
    )
    return {"ok": True, **result}


@router.post("/api/admin/train-bot/actions/promote-qna")
def train_bot_promote_qna(payload: TrainBotPromoteQnaPayload):
    source_item_type = (payload.source_item_type or "").strip()
    if source_item_type not in {"unresolved_queries", "wrong_answer_reports", "user_feedback"}:
        raise HTTPException(status_code=400, detail="Invalid source_item_type")
    result = admin_repository.promote_to_qna_pair(
        source_item_type=source_item_type,
        source_item_id=payload.source_item_id,
        question=payload.question,
        answer=payload.answer,
        category_code=categories_service.normalize(payload.category_code) if payload.category_code else None,
        source_note=payload.source_note,
    )
    return {"ok": True, **result}


@router.post("/api/admin/train-bot/actions/source-reindex")
def train_bot_source_reindex(payload: TrainBotReindexPayload):
    result = admin_repository.trigger_source_reindex_training(payload.data_source_id)
    return {"ok": True, **result}


@router.post("/api/admin/train-bot/actions/category-refresh")
def train_bot_category_refresh(payload: TrainBotCategoryRefreshPayload):
    result = admin_repository.trigger_category_refresh_training(
        categories_service.normalize(payload.category_code) if payload.category_code else None
    )
    return {"ok": True, **result}


@router.post("/api/admin/train-bot/actions/threshold-refresh")
def train_bot_threshold_refresh():
    result = admin_repository.trigger_threshold_refresh_training()
    if not result.get("ok", True):
        return result
    return {"ok": True, **result}


@router.post("/api/admin/train-bot/actions/resolve-wrong-answer")
def train_bot_resolve_wrong_answer(payload: TrainBotResolveWrongAnswerPayload):
    result = admin_repository.resolve_wrong_answer_report(
        report_id=payload.report_id,
        admin_action=payload.admin_action,
        action_notes=payload.action_notes,
    )
    return {"ok": True, **result}


@router.get("/api/admin/categories")
def list_categories(include_inactive: bool = Query(default=True)):
    return {"items": admin_repository.list_categories(include_inactive=include_inactive)}


@router.post("/api/admin/categories")
def create_category(payload: CategoryPayload):
    category_id = admin_repository.create_category(
        code=payload.code,
        name=payload.name,
        description=payload.description,
        display_order=payload.display_order,
        routing_hint=payload.routing_hint,
        prompt_hint=payload.prompt_hint,
        retrieval_scope=payload.retrieval_scope or {},
        is_active=payload.is_active,
    )
    admin_repository.log_admin_action(
        action="create_category_api",
        entity_type="categories",
        entity_id=category_id,
        metadata={"code": payload.code, "name": payload.name},
    )
    return {"ok": True, "category_id": category_id}


@router.put("/api/admin/categories/{category_id}")
def update_category(category_id: int, payload: CategoryPayload):
    ok = admin_repository.update_category(
        category_id,
        {
            "code": payload.code,
            "name": payload.name,
            "description": payload.description,
            "display_order": payload.display_order,
            "is_active": payload.is_active,
            "routing_hint": payload.routing_hint,
            "prompt_hint": payload.prompt_hint,
            "retrieval_scope": payload.retrieval_scope or {},
        },
    )
    if not ok:
        raise HTTPException(status_code=404, detail="Category not found")
    admin_repository.log_admin_action(
        action="update_category_api",
        entity_type="categories",
        entity_id=category_id,
        metadata={"code": payload.code, "name": payload.name},
    )
    return {"ok": True}


@router.post("/api/admin/categories/{category_id}/archive")
def archive_category(category_id: int):
    ok = admin_repository.archive_category(category_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Category not found")
    admin_repository.log_admin_action(
        action="archive_category_api",
        entity_type="categories",
        entity_id=category_id,
        metadata={},
    )
    return {"ok": True}


@router.get("/api/admin/categories/{category_id}/synonyms")
def list_category_synonyms(category_id: int):
    return {"items": admin_repository.list_category_synonyms(category_id)}


@router.post("/api/admin/categories/{category_id}/synonyms")
def add_category_synonym(category_id: int, payload: CategorySynonymPayload):
    try:
        synonym_id = admin_repository.add_category_synonym(category_id, payload.synonym)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {"ok": True, "synonym_id": synonym_id}


@router.get("/api/admin/categories/stats")
def categories_stats():
    return admin_repository.category_statistics()


@router.get("/api/admin/decision-trees")
def list_decision_trees(include_inactive: bool = Query(default=True)):
    return {"items": admin_repository.list_decision_trees(include_inactive=include_inactive)}


@router.get("/api/admin/decision-trees/{tree_id}")
def get_decision_tree(tree_id: int):
    tree = admin_repository.get_decision_tree(tree_id)
    if not tree:
        raise HTTPException(status_code=404, detail="Decision tree not found")
    return {"item": tree}


@router.post("/api/admin/decision-trees")
def save_decision_tree(payload: DecisionTreePayload):
    tree_id = admin_repository.save_decision_tree(payload.model_dump())
    admin_repository.log_admin_action(
        action="save_decision_tree_api",
        entity_type="decision_trees",
        entity_id=tree_id,
        metadata={"name": payload.name},
    )
    return {"ok": True, "tree_id": tree_id}


@router.delete("/api/admin/decision-trees/{tree_id}")
def delete_decision_tree(tree_id: int):
    ok = admin_repository.delete_decision_tree(tree_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Decision tree not found")
    admin_repository.log_admin_action(
        action="delete_decision_tree_api",
        entity_type="decision_trees",
        entity_id=tree_id,
        metadata={},
    )
    return {"ok": True}


@router.get("/api/admin/data-sources")
def list_data_sources():
    return {"items": admin_repository.list_data_sources()}


@router.post("/api/admin/data-sources")
def create_data_source(payload: DataSourceCreatePayload):
    allowed_source_types = {"pdf_folder", "manual_upload", "api", "database", "s3", "gdrive"}
    if payload.source_type not in allowed_source_types:
        raise HTTPException(status_code=400, detail="Unsupported source_type")

    source_id = admin_repository.create_data_source(
        name=payload.name.strip(),
        source_type=payload.source_type,
        source_format=(payload.source_format or "unknown").strip().lower(),
        uri=(payload.uri or "").strip() or None,
    )
    admin_repository.log_admin_action(
        action="create_data_source_api",
        entity_type="data_sources",
        entity_id=source_id,
        metadata={"name": payload.name, "source_type": payload.source_type},
    )
    return {"ok": True, "source_id": source_id}


@router.get("/api/admin/data-sources/{data_source_id}/documents")
def list_source_documents(data_source_id: int):
    return {"items": admin_repository.list_source_documents(data_source_id)}


@router.post("/api/admin/data-sources/{data_source_id}/status")
def set_data_source_status(data_source_id: int, payload: DataSourceStatusPayload):
    if payload.status not in {"enabled", "disabled"}:
        raise HTTPException(status_code=400, detail="Invalid status")
    admin_repository.set_data_source_status(data_source_id, payload.status)
    admin_repository.log_admin_action(
        action="set_data_source_status_api",
        entity_type="data_sources",
        entity_id=data_source_id,
        metadata={"status": payload.status},
    )
    return {"ok": True}


@router.post("/api/admin/data-sources/{data_source_id}/reingest")
def trigger_reingest(data_source_id: int):
    job_id = admin_repository.queue_reingest(data_source_id)
    admin_repository.log_admin_action(
        action="trigger_reingest_api",
        entity_type="data_sources",
        entity_id=data_source_id,
        metadata={"ingestion_job_id": job_id},
    )
    return {"ok": True, "ingestion_job_id": job_id, "status": "queued"}


def _parse_json_records(json_text: str) -> tuple[list[dict], list[str]]:
    errors: list[str] = []
    try:
        payload = json.loads(json_text)
    except json.JSONDecodeError as exc:
        return [], [f"Invalid JSON: {exc.msg} at line {exc.lineno}, column {exc.colno}"]

    if isinstance(payload, dict):
        records = payload.get("records")
        if records is None:
            records = [payload]
    elif isinstance(payload, list):
        records = payload
    else:
        return [], ["JSON must be an object or array of objects"]

    if not isinstance(records, list):
        return [], ["Expected 'records' to be a list"]

    normalized: list[dict] = []
    for i, item in enumerate(records):
        if not isinstance(item, dict):
            errors.append(f"Row {i + 1}: expected object")
            continue
        normalized.append(item)
    return normalized, errors


def _required_fields_for_target(target: str) -> list[str]:
    return {
        "qna_pairs": ["question", "answer"],
        "categories": ["code", "name"],
        "decision_trees": ["name", "nodes"],
        "knowledge_docs": ["title", "content"],
    }.get(target, [])


@router.post("/api/admin/json-convert/preview")
def preview_json_convert(payload: JsonConvertPayload):
    target = (payload.target or "").strip()
    if target not in {"qna_pairs", "categories", "decision_trees", "knowledge_docs"}:
        raise HTTPException(status_code=400, detail="Unsupported target")

    records, parse_errors = _parse_json_records(payload.json_text or "")
    mapping_fields = _required_fields_for_target(target)
    errors = list(parse_errors)

    preview: list[dict] = []
    for idx, record in enumerate(records):
        row_missing = [f for f in mapping_fields if f not in record or record.get(f) in (None, "")]
        if row_missing:
            errors.append(f"Row {idx + 1}: missing required fields: {', '.join(row_missing)}")

        mapped = {f: record.get(f) for f in mapping_fields}
        extra_keys = [k for k in record.keys() if k not in mapping_fields][:8]
        preview.append(
            {
                "row": idx + 1,
                "mapped": mapped,
                "extra_fields": extra_keys,
            }
        )

    return {
        "ok": len(errors) == 0,
        "target": target,
        "record_count": len(records),
        "mapping_fields": mapping_fields,
        "preview": preview[:100],
        "errors": errors,
    }


@router.post("/api/admin/json-convert/import")
def import_json_convert(payload: JsonConvertPayload):
    target = (payload.target or "").strip()
    if target not in {"qna_pairs", "categories", "decision_trees", "knowledge_docs"}:
        raise HTTPException(status_code=400, detail="Unsupported target")

    records, parse_errors = _parse_json_records(payload.json_text or "")
    if parse_errors:
        audit_id = admin_repository.log_import_audit(
            action="json_convert_import",
            target=target,
            status="failed_parse",
            created_count=0,
            error_count=len(parse_errors),
            metadata={"errors": parse_errors},
        )
        return {"ok": False, "created_count": 0, "errors": parse_errors, "audit_log_id": audit_id}

    if target == "qna_pairs":
        created, errors = admin_repository.import_qna_pairs(records)
    elif target == "categories":
        created, errors = admin_repository.import_categories(records)
    elif target == "decision_trees":
        created, errors = admin_repository.import_decision_trees(records)
    else:
        created, errors = admin_repository.import_knowledge_docs(records)

    audit_id = admin_repository.log_import_audit(
        action="json_convert_import",
        target=target,
        status="success" if not errors else "partial",
        created_count=created,
        error_count=len(errors),
        metadata={"record_count": len(records), "errors": errors[:50]},
    )
    response = {
        "ok": len(errors) == 0,
        "target": target,
        "created_count": created,
        "error_count": len(errors),
        "errors": errors,
        "audit_log_id": audit_id,
    }
    admin_repository.log_admin_action(
        action="json_convert_import_api",
        entity_type=target,
        entity_id=audit_id,
        metadata={"created_count": created, "error_count": len(errors)},
    )
    return response


@router.get("/api/admin/qna-pairs")
def list_qna_pairs(
    search: str | None = Query(default=None),
    category: str | None = Query(default=None),
    status: str | None = Query(default="active"),
    approval_status: str | None = Query(default="approved"),
):
    items = admin_repository.list_qna_pairs(
        search=search,
        category_code=categories_service.normalize(category) if category else None,
        status=status,
        approval_status=approval_status,
    )
    return {"items": items}


@router.post("/api/admin/qna-pairs")
def create_qna_pair(payload: QnaPairPayload):
    q = (payload.question or "").strip()
    a = (payload.answer or "").strip()
    if not q or not a:
        raise HTTPException(status_code=400, detail="question and answer are required")

    dupes = admin_repository.duplicate_qna_candidates(q, limit=3)
    pair_id = admin_repository.create_qna_pair(
        question=q,
        answer=a,
        category_code=categories_service.normalize(payload.category_code) if payload.category_code else None,
        source_note=payload.source_note,
        is_exact_eligible=payload.is_exact_eligible,
        is_semantic_eligible=payload.is_semantic_eligible,
        approval_status=payload.approval_status,
        priority=payload.priority,
    )
    admin_repository.log_admin_action(
        action="create_qna_pair_api",
        entity_type="qna_pairs",
        entity_id=pair_id,
        metadata={"category_code": payload.category_code, "priority": payload.priority},
    )
    return {"ok": True, "qna_pair_id": pair_id, "duplicate_candidates": dupes}


@router.put("/api/admin/qna-pairs/{qna_pair_id}")
def update_qna_pair(qna_pair_id: int, payload: QnaPairPayload):
    success = admin_repository.update_qna_pair(
        qna_pair_id,
        {
            "question": payload.question,
            "answer": payload.answer,
            "category_code": categories_service.normalize(payload.category_code) if payload.category_code else None,
            "source_note": payload.source_note,
            "is_exact_eligible": payload.is_exact_eligible,
            "is_semantic_eligible": payload.is_semantic_eligible,
            "approval_status": payload.approval_status,
            "priority": payload.priority,
        },
    )
    if not success:
        raise HTTPException(status_code=404, detail="Q&A pair not found")
    admin_repository.log_admin_action(
        action="update_qna_pair_api",
        entity_type="qna_pairs",
        entity_id=qna_pair_id,
        metadata={"category_code": payload.category_code, "priority": payload.priority},
    )
    return {"ok": True}


@router.post("/api/admin/qna-pairs/{qna_pair_id}/archive")
def archive_qna_pair(qna_pair_id: int):
    success = admin_repository.archive_qna_pair(qna_pair_id)
    if not success:
        raise HTTPException(status_code=404, detail="Q&A pair not found")
    admin_repository.log_admin_action(
        action="archive_qna_pair_api",
        entity_type="qna_pairs",
        entity_id=qna_pair_id,
        metadata={},
    )
    return {"ok": True}


@router.delete("/api/admin/qna-pairs/{qna_pair_id}")
def delete_qna_pair(qna_pair_id: int):
    success = admin_repository.delete_qna_pair(qna_pair_id)
    if not success:
        raise HTTPException(status_code=404, detail="Q&A pair not found")
    admin_repository.log_admin_action(
        action="delete_qna_pair_api",
        entity_type="qna_pairs",
        entity_id=qna_pair_id,
        metadata={},
    )
    return {"ok": True}


@router.get("/api/admin/qna-pairs/duplicates")
def qna_duplicates(question: str = Query(..., min_length=1)):
    return {"items": admin_repository.duplicate_qna_candidates(question, limit=5)}


@router.post("/api/admin/expert-answer")
def save_expert_answer(payload: ExpertAnswerPayload):
    category = categories_service.normalize(payload.category)

    if not category:
        return {
            "ok": False,
            "message": "Invalid category",
        }

    normalized_question = normalize_question_text(
        payload.normalized_question or payload.question
    )

    expert_answer_id = expert_answers_service.save(
        question=payload.question,
        normalized_question=normalized_question,
        category=category,
        expert_answer=payload.expert_answer,
        source_note=payload.source_note,
        unresolved_query_id=payload.unresolved_query_id,
    )

    return {
        "ok": True,
        "expert_answer_id": expert_answer_id,
    }


@router.get("/api/admin/export/unresolved")
def export_unresolved(
    category: str | None = Query(default=None),
    status: str = Query(default="open"),
):
    norm = categories_service.normalize(category) if category else None
    items = admin_repository.list_unresolved(norm, status)

    wb = Workbook()
    ws = wb.active
    ws.title = "Unresolved Queries"

    headers = ["ID", "Category", "Question", "Reason", "Created"]
    ws.append(headers)

    for item in items:
        ws.append([
            item.get("id"),
            item.get("final_category") or item.get("category"),
            item.get("question"),
            item.get("reason"),
            item.get("created_at"),
        ])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=unresolved_queries.xlsx"
        },
    )


@router.get("/api/admin/export/feedback")
def export_feedback(category: str | None = Query(default=None)):
    norm = categories_service.normalize(category) if category else None
    items = admin_repository.list_feedback(norm)

    wb = Workbook()
    ws = wb.active
    ws.title = "User Feedback"

    headers = ["ID", "Category", "Question", "Status", "Comment", "Created"]
    ws.append(headers)

    for item in items:
        ws.append([
            item.get("id"),
            item.get("category"),
            item.get("question"),
            "Satisfied" if item.get("satisfied") else "Not satisfied",
            item.get("comment"),
            item.get("created_at"),
        ])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=user_feedback.xlsx"
        },
    )


@router.get("/api/admin/export/chat-history")
def export_chat_history(
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    category: str | None = Query(default=None),
    response_mode: str | None = Query(default=None),
    feedback_status: str | None = Query(default=None),
):
    items = admin_repository.list_chat_sessions(
        date_from=date_from,
        date_to=date_to,
        category_code=categories_service.normalize(category) if category else None,
        response_mode=response_mode,
        feedback_status=feedback_status,
        limit=1000,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Chat Sessions"
    ws.append([
        "Session ID",
        "Session Key",
        "Status",
        "Started At",
        "Ended At",
        "Messages",
        "Answers",
        "Feedback",
        "Unsatisfied",
        "Wrong Reports Open",
        "Admin Note",
    ])
    for item in items:
        ws.append([
            item.get("id"),
            item.get("session_key"),
            item.get("status"),
            item.get("started_at"),
            item.get("ended_at"),
            item.get("message_count"),
            item.get("answer_count"),
            item.get("feedback_count"),
            item.get("unsatisfied_count"),
            item.get("wrong_answer_open_count"),
            item.get("admin_note"),
        ])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=chat_history_sessions.xlsx"},
    )
